from playwright.sync_api import sync_playwright
import time
from to_exel import ExcelManager

import openpyxl


def fill_empty_cells_with_studio(file_path, sheet_name, column_name):
    # Загружаем файл Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Получаем номер столбца по его имени
    column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == column_name:
            column = col
            break

    if column is None:
        print(f"Столбец {column_name} не найден.")
        return

    # Перебираем ячейки в столбце, начиная со второй строки
    for cell in sheet.iter_cols(min_col=column, max_col=column):
        for c in cell:
            if c.row == 1:
                continue  # Пропускаем заголовок
            if c.value is None:
                # Если значение ячейки пустое, заменяем его на "studio"
                c.value = "studio"

    # Сохраняем изменения
    wb.save(file_path)


data = []
manager = ExcelManager('test_1.xlsx', ['title', 'location', 'city', 'room', 'beth', 'area', 'price'])


def main():
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        page = browser.new_page()
        for _ in range(1, 307):
            try:
                url = f'https://dubai.dubizzle.com/property-for-sale/residential/?neighborhood=57505&neighborhood=193&neighborhood=52&neighborhood=87&neighborhood=63&page={_}'
                page.goto(url, timeout=900000)

                page.wait_for_selector('div[data-testid="listing-price"]', timeout=900000)
                prices = page.query_selector_all('div[data-testid="listing-price"]')
                titles = page.query_selector_all('h2[data-testid="subheading-text"]')
                rooms = page.query_selector_all(
                    '//div[contains(@class, "property-text") and @data-testid="listing-bedrooms"]')
                baths = page.query_selector_all(
                    '//div[contains(@class, "property-text") and @data-testid="listing-bathrooms"]')
                areas = page.query_selector_all('//div[contains(@class, "property-text") and @data-testid="listing-size"]')
                locations = page.query_selector_all('//div[@data-testid="listing-location"]/span')
                for price, title, room, bath, area, location in zip(prices, titles, rooms, baths, areas, locations):
                    city = location.inner_text().split(', ')[-1]
                    data.append([
                        title.text_content(),
                        location.text_content(),
                        city,
                        room.text_content(),
                        bath.text_content(),
                        area.text_content(),
                        price.text_content()

                    ])
            except:
                continue
        manager.save_to_excel(data)


#main()

fill_empty_cells_with_studio("test_1.xlsx", "List_1", "room")
